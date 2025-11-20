import os
import pandas as pd
from openpyxl import load_workbook

# =========================
# Paths
# =========================
working_folder = r'C:\Users\PAWARUX1\Manager working summary'
input_file = 'AIL_Effort KPI_YTD 2025.xlsb'
report_file = 'report.csv'
tbm_file = 'TBM.csv'
output_file = 'ABM_Consolidated_Report.xlsx'

input_filepath = os.path.join(working_folder, input_file)
report_filepath = os.path.join(working_folder, report_file)
tbm_filepath = os.path.join(working_folder, tbm_file)

report_output_folder = os.path.join(working_folder, "ABM_Reports")
template_path = os.path.join(working_folder, "Report Template.xlsx")

os.makedirs(report_output_folder, exist_ok=True)

# =========================
# STEP 1: Read ABM data
# =========================
print("Reading ABM data...")
df_abm = pd.read_excel(input_filepath, sheet_name='ABM', engine='pyxlsb')

# Identify ABM Name column
abm_name_col = None
for col in df_abm.columns:
    col_str = str(col).upper().replace(' ', '-').replace('_', '-')
    if 'ABM' in col_str and 'NAME' in col_str:
        abm_name_col = col
        break
if not abm_name_col:
    raise ValueError("Could not find ABM Name column")

# Clean ABM Name column
df_abm = df_abm[df_abm[abm_name_col].notna()]
df_abm = df_abm[df_abm[abm_name_col].astype(str).str.strip() != '']
df_abm.rename(columns={abm_name_col: 'ABM Name'}, inplace=True)

# Normalize ABM Name and Month Start for matching
df_abm['ABM Name'] = df_abm['ABM Name'].astype(str).str.strip().str.lower()
if 'Month Start' not in df_abm.columns:
    raise ValueError("ABM data must contain 'Month Start' column")
df_abm['Month Start'] = df_abm['Month Start'].astype(str).str.strip()

# =========================
# STEP 2: Read report.csv and TBM.csv
# =========================
print("Reading report.csv and TBM.csv...")
try:
    report_df = pd.read_csv(report_filepath, encoding='utf-8')
except UnicodeDecodeError:
    report_df = pd.read_csv(report_filepath, encoding='ISO-8859-1')

try:
    tbm_df = pd.read_csv(tbm_filepath, encoding='utf-8')
except UnicodeDecodeError:
    tbm_df = pd.read_csv(tbm_filepath, encoding='ISO-8859-1')

# Extract TBM names list
tbm_names = tbm_df.iloc[:, 0].dropna().astype(str).tolist()
tbm_names_lower = [tbm.strip().lower() for tbm in tbm_names]

# Validate report.csv columns
required_cols = ['User: Full Name', 'Work With', 'Date']
for col in required_cols:
    if col not in report_df.columns:
        raise ValueError(f"report.csv must contain '{col}' column")

# Normalize names and prepare month column
report_df['User: Full Name'] = report_df['User: Full Name'].astype(str).str.strip().str.lower()

# Convert Date column to datetime using dd-mm-yy format
report_df['Date'] = pd.to_datetime(report_df['Date'], format='%d-%m-%y', errors='coerce')

# Extract month-year in format Jan'25
report_df['Month'] = report_df['Date'].dt.strftime("%b'%y")  # e.g., Jan'25

# =========================
# STEP 3: Reliable TBM count per row
# =========================
def count_tbms_for_row(abm_name, month):
    # Filter report for ABM and month
    abm_rows = report_df[(report_df['User: Full Name'] == abm_name) &
                          (report_df['Month'] == month)]
    tbm_worked = set()
    for val in abm_rows['Work With'].dropna():
        # Split by comma, normalize each name
        names = [name.strip().lower() for name in str(val).split(',') if name.strip()]
        for name in names:
            if name in tbm_names_lower:
                tbm_worked.add(name)
    return len(tbm_worked)

# Apply logic row-wise
df_abm['No. of TBMs Worked'] = df_abm.apply(
    lambda row: count_tbms_for_row(row['ABM Name'], row['Month Start']),
    axis=1
)

# =========================
# STEP 4: Rename headers and save consolidated summary
# =========================
rename_headers = {
    "Call-Days": "No. of Days",
    "Actual-DR-Calls": "No. of Calls",
    "2PC-Freq-Cov-%": "2 PC %Coverage",
    "Total-DR-Cov-%": "%Total Coverage"
}

df_abm.rename(columns={col: rename_headers[col] for col in rename_headers if col in df_abm.columns}, inplace=True)

print("Saving consolidated summary...")
df_abm.to_excel(os.path.join(working_folder, output_file), index=False)
print(f"✓ Consolidated summary saved to: {output_file}")

# =========================
# STEP 5: Generate reports using template from consolidated file
# =========================
print("Generating reports using template...")

# Row mapping based on template
row_map = {
    "No. of Days": 2,
    "No. of Calls": 3,
    "No. of TBMs Worked": 4,
    "No. of HQs Visited": 5,
    "%Total Coverage": 8,
    "2 PC %Coverage": 9,
    "Call Average": 10
}
nov_col = 3  # Column C for Nov
name_cell = "A1"

# Fields mapping from consolidated file
report_fields = {
    "No. of Days": "No. of Days",
    "No. of Calls": "No. of Calls",
    "No. of TBMs Worked": "No. of TBMs Worked",
    "No. of HQs Visited": "No. of TBMs Worked",  # Same as TBMs Worked
    "%Total Coverage": "%Total Coverage",
    "2 PC %Coverage": "2 PC %Coverage"
}

# Generate reports for each ABM from consolidated file
for abm_name, group_df in df_abm.groupby('ABM Name'):
    values = {}
    for metric, col in report_fields.items():
        if col in group_df.columns:
            if metric in ["%Total Coverage", "2 PC %Coverage"]:
                values[metric] = round(group_df[col].mean(skipna=True), 2)
            else:
                values[metric] = int(group_df[col].sum(skipna=True))

    # Compute Call Average
    no_days = values.get("No. of Days", 0)
    no_calls = values.get("No. of Calls", 0)
    if no_days or no_calls:
        values["Call Average"] = round((no_days + no_calls) / 2, 2)
    else:
        values["Call Average"] = ""

    # Load template
    wb = load_workbook(template_path)
    ws = wb.active

    # Insert ABM name
    ws[name_cell] = f"ABM: {abm_name}"

    # Fill Nov column values
    for metric, row in row_map.items():
        ws.cell(row=row, column=nov_col).value = values.get(metric, "")

    # Save as new report file
    report_file_path = os.path.join(report_output_folder, f"{abm_name}_report.xlsx")
    wb.save(report_file_path)

print("✅ Reports created using template with correct layout and month-wise TBM values.")
print("\n🎯 END-TO-END PROCESS COMPLETE!")